# *- coding: utf-8 -*-
import os
from tqdm import tqdm
import shutil
import requests
from openpyxl.reader.excel import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor


def downloader(path: str):
    # URL = "https://images.puma.net/images/360248/08/fnd/ALL"
    # image_path = "res.jpg"
    width = 150
    height = 150
    xlsx_path = path
    wb = load_workbook(xlsx_path)
    # wb = Workbook()
    ws = wb.active

    rows_quant = ws.max_row - 1
    num = 3

    if os.path.isdir("images"):
        shutil.rmtree("images")

    os.mkdir("images")

    for i in tqdm(range(rows_quant), desc="загружаю картинки"):
        file_path = f"images/res{i}.jpg"
        URL = ws.cell(row=i + 2, column=3).hyperlink.target
        # print(URL)
        # print(i)
        r = requests.get(URL, stream=True)

        # print(r.status_code)
        if r.status_code == 200:
            with open(file_path, 'wb') as f:
                r.raw.decode_content = True
                shutil.copyfileobj(r.raw, f)

        ws.column_dimensions['P'].width = 20.0
        ws.row_dimensions[i + 2].height = 116

        col, row = 15, i + 1
        offset = 30000
        img = Image(file_path)
        _from = AnchorMarker(
            col=col,
            row=row,
            colOff=offset,
            rowOff=offset,
        )
        to = AnchorMarker(
            col=col + 1,
            row=row + 1,
            colOff=-offset,
            rowOff=-offset,
        )

        img.anchor = TwoCellAnchor(editAs="twoCell", _from=_from, to=to)

        cell = f"P{row + 1}"
        ws.add_image(img)

    print("гружу в excel...ждём")

    wb.save(xlsx_path)

    # os.rmdir("images")
    shutil.rmtree("images")
